import openpyxl
from openpyxl.styles import Font
BASE="/mnt/data2/墨江紫米研究"
OLD=f"{BASE}/12_论文Paper1_Manuscript/supplementary/Supplementary_Tables.xlsx"
OUT=f"{BASE}/12_论文Paper1_Manuscript/_SUBMISSION_PlantComm/Supplementary_Tables.xlsx"
ANA=f"{BASE}/12_论文Paper1_Manuscript/_SUBMISSION_PlantComm/analysis"
old=openpyxl.load_workbook(OLD, data_only=True)
new=openpyxl.Workbook(); new.remove(new.active)
bold=Font(bold=True)

def copy_sheet(old_name, new_name):
    o=old[old_name]; s=new.create_sheet(new_name)
    for r,row in enumerate(o.iter_rows(values_only=True),1):
        for c,v in enumerate(row,1):
            if v is not None: s.cell(r,c,v)
    for c in range(1,o.max_column+1): s.cell(1,c).font=bold
    return s

def write_rows(name, rows):
    s=new.create_sheet(name)
    for r,row in enumerate(rows,1):
        for c,v in enumerate(row,1):
            if v is not None: s.cell(r,c,v)
    for c in range(1,len(rows[0])+1): s.cell(1,c).font=bold
    return s

def read_tsv(path):
    return [ln.rstrip("\n").split("\t") for ln in open(path)]

# S1 assembly (copy + append spanning-read validation)
s1=copy_sheet("S1_Assembly","S1_Assembly_QC")
extra=[[None],["Long-read validation of OsB2 insertions","",""],
 ["HiFi min depth across Kala4/OsB2 locus","≥11x",""],
 ["HiFi reads spanning promoter TE cluster","31",""],
 ["HiFi reads spanning intronic 12.5-kb Gypsy","10",""],
 ["Read-mapping rate (HiFi / ONT / Illumina)","99.96% / 99.99% / 99.31%",""]]
r0=s1.max_row+1
for i,row in enumerate(extra):
    for c,v in enumerate(row,1):
        if v is not None: s1.cell(r0+i,c,v)

# S2 annotation (copy)
copy_sheet("S2_Annotation","S2_Annotation")

# S3 SV (authoritative syri.summary aggregate + asm sensitivity)
s3=[["Table S3. Genome-wide structural variation, ZN65 vs Nipponbare (minimap2 asm5 + SyRI). Structural counts are SyRI per-event annotations; the aggregate SyRI summary groups translocations+inverted translocations and duplications+inverted duplications.","",""],
 ["Variant class","Count","Length (ref/qry, bp)"],
 ["Sequence-level","",""],
 ["  SNP","1,045,956","1,045,956"],
 ["  Insertion","88,805","2,890,667 (qry)"],
 ["  Deletion","91,649","2,447,558 (ref)"],
 ["Structural","",""],
 ["  Inversion","121","8,060,425 / 10,089,103"],
 ["  Translocation (aggregate)","449","= 198 TRANS + 251 inverted-TRANS"],
 ["  Duplication (reference-side)","157","1,030,576"],
 ["  Duplication (query-side)","632","2,701,181"],
 ["  Highly diverged region","7,466","149,775,151 / 162,381,462"],
 ["  Not aligned (ref / qry)","1,149 / 1,536","63,702,354 / 70,339,400"],
 ["Syntenic regions","527","298,992,637 / 310,054,998"],
 [None],
 ["Robustness to minimap2 stringency (Inversions; ref-side Duplications; Translocations)","",""],
 ["asm5","121 INV; 157 DUP; 449 TRANS",""],
 ["asm10","132 INV; 133 DUP; 497 TRANS",""],
 ["asm20","126 INV; 62 DUP; 317 TRANS",""],
 ["Orthogonal validation","nucmer/Assemblytics (concordant)",""]]
write_rows("S3_SV_genomewide", s3)

# S4 pathway copy number (new OG table) + full inventory appended
pw=read_tsv(f"{ANA}/pathway_copynumber_ZN65_vs_NIP.tsv")
s4=[["Table S4. Flavonoid/anthocyanin pathway copy number, ZN65 vs Nipponbare (KEGG orthology + OrthoFinder orthogroup baseline). Orthogroup counts are distinct genes in the orthogroups spanning each step and may include paralogues; ZN65 shows no pathway-wide expansion.","","","",""]]
s4.append(pw[0])
for r in pw[1:]: s4.append(r)
s4.append([None]); s4.append(["Full 100-gene ZN65 pathway inventory (KEGG orthology)","","","",""])
oi=old["S4_Pathway_genes"]
for row in oi.iter_rows(values_only=True):
    if any(v is not None for v in row): s4.append([("" if v is None else v) for v in row])
write_rows("S4_Pathway_copynumber", s4)

# S5 centromere + rDNA (new)
cen=read_tsv(f"{ANA}/centromere_table.tsv"); rdn=read_tsv(f"{ANA}/rDNA_table.tsv")
s5=[["Table S5. Centromere (CentO satellite) and 45S/5S ribosomal DNA arrays localized in the ZN65 T2T assembly.","","","","",""],
 ["Centromeres (CentO 155-165 bp tandem-satellite arrays)","","","","",""]]
s5.append(cen[0])
for r in cen[1:]: s5.append(r)
s5.append([None]); s5.append(["Ribosomal DNA arrays","","","","",""]); s5.append(rdn[0])
for r in rdn[1:]: s5.append(r)
s5.append([None]); s5.append(["Note: 45S NOR = major array on Chr9 short arm (~720 units); minor Chr10 cluster. 5S array on Chr11 (~1,085 units). Consistent with the canonical rice rDNA architecture.","","","","",""])
write_rows("S5_Centromere_rDNA", s5)

# S6 OsB2 transposons (copy)
copy_sheet("S6_OsB2_transposons","S6_OsB2_transposons")
# S7 pigmentation-loci alleles (from old S5)
copy_sheet("S5_Pigmentation_loci","S7_Pigmentation_loci_alleles")
# S8 nomenclature (copy)
copy_sheet("S8_Nomenclature","S8_Nomenclature")
# S9 Os02g (from old S7)
copy_sheet("S7_Os02g_identification","S9_Os02g_identification")
# S10 P/A (copy)
copy_sheet("S10_OsB2_insertion_PA","S10_OsB2_insertion_PA")
# S11 subspecies (copy)
copy_sheet("S11_ZN65_subspecies","S11_ZN65_subspecies")

# S12 expression (new)
ke=read_tsv(f"{ANA}/expression_key_TPM.tsv")
s12=[["Table S12. Pigmentation/MBW gene expression in vegetative tissue (TPM). Seedling SD n=3, tillering TG n=2 (TG2 absent); no biological replicate for one contrast, so no inferential differential-expression testing was performed. Not pericarp tissue.","","","","","","","","",""]]
s12.append(ke[0])
for r in ke[1:]: s12.append(r)
s12.append([None]); s12.append(["Pattern: OsC1/Kala3, Rc and OsDFR/Kala1 essentially silent (<1 TPM); OsTTG1 constitutive; OsB2 partial/variable. 64% of the 100 pathway genes >=1 TPM.","","","","","","","","",""])
write_rows("S12_Expression_TPM", s12)

# S13 polyphenol (from old S9)
copy_sheet("S9_TotalPolyphenol_interim","S13_TotalPolyphenol_interim")

new.save(OUT)
print("Wrote", OUT)
print("Sheets:", new.sheetnames)
